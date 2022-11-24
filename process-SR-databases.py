"""
A program for **making** MÁV's internal, really **dirty database** of speed restrictions **computer-processable.**

Receives:
several _speed restriction databases_ in Excel formats;
a .csv _database of stations,_ incl. their names and UIC codes;
a .csv _database of UIC member railway companies,_ incl. their names and UIC codes;
a .csv _database of UIC member countries,_ incl. their names, ISO and UIC codes.

Returns a neat, computer-processable _.csv database of all speed restrictions_ with extended contents.

Abbreviations:
_SR = speed restriction_;
_TSR = temporary speed restriction_;
_ASR = all speed restrictions_.
"""

from openpyxl import load_workbook
import pandas as pd
import re  # built-in
import roman
import numpy as np
import csv  # built-in
from datetime import datetime
import zoneinfo
from natsort import index_realsorted


def import_csv(file_location, index_col_loc):
    """
    Receives a .csv file's location and the index column's number.
    Returns the .csv file's content as a `pandas` `DataFrame`.
    """
    print(file_location + ' will now be imported!')
    return pd.read_csv(file_location, sep=';', index_col=index_col_loc, dtype=str)


def import_worksheets(excel_file_location):
    """
    Receives an _Excel file location_ and a _database_.

    Imports an _Excel file_ using `openpyxl`.
    Appends _all worksheets_ from the _Excel file_ to the _database_.

    Returns the _database_.
    """
    excel_workbook = load_workbook(excel_file_location)
    print(excel_file_location + ' loaded!')

    current_sr_file_openpyxl = []
    for excel_worksheet in excel_workbook.worksheets:
        current_sr_file_openpyxl.append(excel_worksheet)
    print('All worksheets imported from ' + excel_file_location + ' to current_sr_file_openpyxl!')
    return current_sr_file_openpyxl


def check_permanence(cell):
    """
    Receives an Excel file's _cell_ imported via `openpyxl`.
    Returns _if that cell is bold_ using `openpyxl`.
    """
    return cell.font.bold


def should_be_imported(import_all, is_it_temporary):
    """
    Checks if the current SR to import should already have been imported, therefore if it should be done now.
    """
    return import_all is True or is_it_temporary is False


def convert_openpyxl_cells_to_values(row):
    """
    Receives _cells_ of an Excel file's row imported via `openpyxl`.
    Returns the same row but with just _each cell's values_.
    """
    for idx_02, cell in enumerate(row):
        row[idx_02] = cell.value


class SpeedRestriction:
    """
    Stores all data of a speed restriction.
    """

    def __init__(self,
                 country_iso=None,
                 country_uic=None,
                 company_short_name=None,
                 company_uic=None,
                 line=None,
                 station_from_name=None,
                 station_from_uic=None,
                 station_to_name=None,
                 station_to_uic=None,
                 on_station=None,
                 track_switch_source_text=None,
                 applied_to=None,
                 number_of_open_line_tracks=None,
                 track_open_line=None,
                 track_station=None,
                 switch_station=None,
                 in_timetable=None,
                 location_from=None,
                 location_to=None,
                 length=None,
                 speed=None,
                 length_time=None,
                 internal_id=None,
                 time_from=None,
                 cause=None,
                 time_to=None,
                 comment=None):
        self.country_iso = country_iso
        self.country_uic = country_uic
        self.company_short_name = company_short_name
        self.company_uic = company_uic
        self.line = line
        self.station_from_name = station_from_name
        self.station_from_uic = station_from_uic
        self.station_to_name = station_to_name
        self.station_to_uic = station_to_uic
        self.on_station = on_station
        self.track_switch_source_text = track_switch_source_text
        self.applied_to = applied_to
        self.number_of_open_line_tracks = number_of_open_line_tracks
        self.track_open_line = track_open_line
        self.track_station = track_station
        self.switch_station = switch_station
        self.in_timetable = in_timetable
        self.location_from = location_from
        self.location_to = location_to
        self.length = length
        self.speed = speed
        self.length_time = length_time
        self.internal_id = internal_id
        self.time_from = time_from
        self.cause = cause
        self.time_to = time_to
        self.comment = comment

    def __iter__(self):
        return iter([self.country_iso,
                     self.country_uic,
                     self.company_short_name,
                     self.company_uic,
                     self.line,
                     self.station_from_name,
                     self.station_from_uic,
                     self.station_to_name,
                     self.station_to_uic,
                     self.on_station,
                     self.track_switch_source_text,
                     self.applied_to,
                     self.number_of_open_line_tracks,
                     self.track_open_line,
                     self.track_station,
                     self.switch_station,
                     self.in_timetable,
                     self.location_from,
                     self.location_to,
                     self.length,
                     self.speed,
                     self.length_time,
                     self.internal_id,
                     self.time_from,
                     self.cause,
                     self.time_to,
                     self.comment])


def add_data(row,
             sr_database,
             in_timetable,
             country_uic_database,
             company_uic_database,
             station_database):
    """
    Receives a _SR's row,_
    a _database_,
    if the SR is _in the timetables,_
    the .csv _database of UIC member countries,_
    the .csv _database of UIC member railway companies,_
    the .csv _database of stations,_.

    Processes all data in the row.

    Appends the _processed row_ to the _database._
    """

    def lookup_data(what, where, in_which_column):
        """
        Receives a _text,_ a _database,_ and a _column name_ to look for.
        Returns the _text_ in the same row's searched column where the input text was _found._
        Raises an error if searched data could not be found.
        """
        try:
            return str(where.at[what, in_which_column])
        except KeyError:
            print('Error: ' + what + ' could not be converted to ' + in_which_column + '!')
            return 'KeyError'

    def remove_space_after_hyphen(data):
        return re.sub(r'(?<=\w)- (?=\w)', '-', str(data))

    def station_name_search(regex, text):
        try:
            return re.search(regex, text)[0]
        except TypeError:
            return 'TypeError'

    def extract_number(data):
        """
        Receives a text.

        Initializes *regex search expressions* for arabic and roman numbers with several combinations
        that could be found in the database.

        Searches for an _arabic number._
        If found, returns it.
        If not found, searches for a _roman number._

        If found, converts it to an _arabic number_ via `roman` and returns it.
        If this doesn't succeed, returns an `InvalidRomanNumeralError`.
        If not found, searches for an _arabic number with letters._

        If found, returns it.
        If not found, returns a `TypeError`.
        """
        text = str(data)

        arabic_number_regex = r'((?:^)|(?<= |\.|\())' \
                              r'\d+' \
                              r'(?=\D)'
        find_roman_number_regex = r'((?:^)|(?<= |\.|\())' \
                                  r'((M{0,4}(CM|CD|D?C{0,3})(XC|XL|L?X{0,3})(IX|IV|V?I{0,3}))' \
                                  r'|(M{0,4}(CM|CD|D?C{0,3})(XC|XL|L?X{0,3})(IX|IV|V?I{0,3}))\/[a-z]' \
                                  r'|(M{0,4}(CM|CD|D?C{0,3})(XC|XL|L?X{0,3})(IX|IV|V?I{0,3}))[a-z])' \
                                  r'(?=\.| )'
        get_roman_number_regex = r'M{0,4}(CM|CD|D?C{0,3})(XC|XL|L?X{0,3})(IX|IV|V?I{0,3})'
        letter_and_arabic_number_regex = r'((?:^)|(?<= |\.|\())\w\d+(?=\D)'

        # TODO: InvalidRomanNumeralError:
        #  "Chinoin" vontató vágány és a kapcsolódó
        #  VII/a. sz. vágány
        #  XIVa.vg.
        #  B II/a. vágány

        # TODO: TypeError:
        #  "T" vágány
        #  AUDI kihúzóvágány
        #  Duna vontatóvágány
        #  "Amerikai" kihúzó vágány
        #  "Újkezelő" vágány
        #  Józsefvárosi összekötő vágány
        #  "Házi" csonka vágány
        #  Homlokrakodó csonka vágány
        #  MOL kihúzó
        #  Fűtőház vágány

        try:
            if re.search(arabic_number_regex, text, re.MULTILINE) is not None:
                return re.search(arabic_number_regex, text, re.MULTILINE)[0]
            else:
                if re.search(find_roman_number_regex, text, re.MULTILINE) is not None:
                    if re.search(find_roman_number_regex, text, re.MULTILINE)[0] != '':
                        roman_number = re.search(find_roman_number_regex, text, re.MULTILINE)[0]
                        return roman.fromRoman(roman_number)
                    else:
                        return re.search(letter_and_arabic_number_regex, text, re.MULTILINE)[0]
                else:
                    return re.search(letter_and_arabic_number_regex, text, re.MULTILINE)[0]
        except TypeError:
            return 'TypeError'
        except roman.InvalidRomanNumeralError:
            return 'InvalidRomanNumeralError'

    def convert_date_to_iso(text):
        try:
            return \
                datetime.strptime(text, '%Y.%m.%d %H:%M')\
                .replace(tzinfo=zoneinfo.ZoneInfo(key='Europe/Budapest'))\
                .isoformat()
        except TypeError:
            return 'TypeError'
        except ValueError:
            return 'ValueError'

    current_sr = SpeedRestriction()

    current_sr.country_iso = 'HU'
    current_sr.country_uic = lookup_data(what=current_sr.country_iso,
                                         where=country_uic_database,
                                         in_which_column='Numerical code')
    current_sr.company_short_name = 'MÁV'
    current_sr.company_uic = lookup_data(what=current_sr.company_short_name,
                                         where=company_uic_database,
                                         in_which_column='code')

    current_sr.line = row[0]
    current_sr.station_from_name = remove_space_after_hyphen(row[1])
    current_sr.station_from_uic = station_name_search(r'(?<=HU).*',
                                                      lookup_data(what=current_sr.station_from_name,
                                                                  where=station_database,
                                                                  in_which_column='PLC kód'))

    current_sr.on_station = not row[2]  # true if cell is empty
    if current_sr.on_station:
        current_sr.track_switch_source_text = row[4]
        if str(row[4]).find('kitérő') != -1:
            current_sr.applied_to = 'switch'
            current_sr.switch_station = extract_number(row[4])
        else:
            current_sr.applied_to = 'track'
            current_sr.track_station = extract_number(row[4])
    else:
        current_sr.station_to_name = remove_space_after_hyphen(row[2])
        current_sr.station_to_uic = station_name_search(r'(?<=HU).*',
                                                        lookup_data(what=current_sr.station_to_name,
                                                                    where=station_database,
                                                                    in_which_column='PLC kód'))
        current_sr.applied_to = 'track'
        if row[3] is None:
            current_sr.number_of_open_line_tracks = 1
        elif row[3] == 'bal vágány':
            current_sr.number_of_open_line_tracks = 2
            current_sr.track_open_line = 'left'
        elif row[3] == 'jobb vágány':
            current_sr.number_of_open_line_tracks = 2
            current_sr.track_open_line = 'right'
        else:
            current_sr.number_of_open_line_tracks = current_sr.track_open_line = 'UnknownError'

    # TODO: save roman numerals as well
    # TODO: track's UIC code
    # TODO: SR's direction
    # TODO: SR's permanence

    current_sr.in_timetable = in_timetable
    current_sr.location_from = row[5]
    current_sr.location_to = row[6]

    current_sr.internal_id = row[10]
    current_sr.time_from = convert_date_to_iso(row[11])
    current_sr.cause = row[12]
    current_sr.time_to = convert_date_to_iso(row[13]) if row[13] else None
    current_sr.comment = row[14]

    # TODO: megszüntetés (tervezett) dátuma

    sr_database.append(current_sr)


def edit_full_df(full_df):
    """
    Removes leftover rows from the Excel import.

    Sorts the database.
    """
    full_df.query("line != 'Vonal'", inplace=True)
    full_df.query("location_from != 'Hossz (m):'", inplace=True)
    full_df['line'] = full_df['line'].astype(str)
    full_df.sort_values(by=['line', 'location_from'],
                        key=lambda x: np.argsort(
                            index_realsorted(
                                zip(full_df['line'],
                                    full_df['location_from']))),
                        inplace=True)
    print('full_df edited!')


def main():
    """
    Initializes external database locations.

    Imports SR databases via `import_worksheets`.
    Creates a list database from each file via `create_database`,
    taking in account if all SRs should be imported from them.
    Converts the list database to a dataframe via `create_dataframe`.
    Edits the dataframe via `edit_full_df`.
    Saves the dataframe via `save_file`.
    """
    tsr_excel_file_location = 'data/SR_lists/MÁV_220202_TSR.xlsx'
    asr_excel_file_location = 'data/SR_lists/MÁV_220202_ASR.xlsx'
    sr_file_locations = [tsr_excel_file_location, asr_excel_file_location]
    sr_database = []
    print('Variables initialized!')

    country_uic_database = import_csv('data/UIC_lists/country_UIC_list.csv', 1)
    company_uic_database = import_csv('data/UIC_lists/company_UIC_list.csv', 1)
    station_database = import_csv('data/station_lists/stations_HU.csv', 0)

    for i, file_location in enumerate(sr_file_locations):
        current_sr_file_openpyxl = import_worksheets(file_location)
        if i == 0:  # handle duplicated SRs
            import_all = True
        else:
            import_all = False
        for excel_worksheet in current_sr_file_openpyxl:
            for row in excel_worksheet.iter_rows(min_row=2):
                row = list(row)  # convert from tuples
                is_it_temporary = check_permanence(row[0])
                if should_be_imported(import_all, is_it_temporary):
                    convert_openpyxl_cells_to_values(row)
                    add_data(row,
                             sr_database,
                             is_it_temporary,
                             country_uic_database,
                             company_uic_database,
                             station_database)

    with open('export/_all_220202_ASR_02.csv', mode='w+') as file:
        writer = csv.writer(file)
        for sr in sr_database:
            writer.writerow(list(sr))
    print('sr_list saved to file!')

    print('All done!')


main()
